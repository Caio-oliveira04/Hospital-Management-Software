import openpyxl
import pathlib
import os
import time

def clear_screen():
    os.system('cls' if os.name == 'nt' else 'clear')

class User:
    def __init__(self):
        self.name = ""
        self.age = None
        self.email = ""
        self.password = ""
        self.phone = ""

    def set_name(self, name):
        self.name = name

    def get_name(self):
        return self.name

    def set_age(self, age):
        self.age = age

    def get_age(self):
        return self.age

    def set_email(self, email):
        self.email = email

    def get_email(self):
        return self.email

    def set_password(self, password):
        self.password = password

    def get_password(self):
        return self.password

    def set_phone(self, phone):
        self.phone = phone

    def get_phone(self):
        return self.phone

    def register_user(self, name, email, password):
        try: 
            file_path = pathlib.Path("users.xlsx")

            if file_path.exists():
                workbook = openpyxl.load_workbook("users.xlsx")
            else:
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet['A1'] = "Name"
                sheet['B1'] = "Email"
                sheet['C1'] = "Password"
                sheet

            sheet = workbook.active
            sheet.append([name, email, password])

            workbook.save("users.xlsx")
            clear_screen()  
            print('Registration successful')
            time.sleep(2)  

        except FileNotFoundError:
            print('File not found. Make sure the users.xlsx file exists.')

        except Exception as e:
            print(f'Error saving data: {e}')

    def login_user(self, email, password):
        try:
            file_path = pathlib.Path("users.xlsx")
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active

            for row in sheet.iter_rows(values_only=True):
                if row[1] == email and row[2] == password:
                    print("Login successful!")
                    return
                
                elif row[1] == email and row[2] != password:
                    print("Incorrect password")
                    return

                elif row[1] != email and row[2] == password:
                    print("Incorrect email")
                    return

            print("Email not found.")

        except FileNotFoundError:
            print(f'File not found: {file_path}')
        except Exception as e:
            print(f'Error during login: {e}')

    def schedule_appointment(self):
        try:
            specialist = int(input("Select the specialist:\n"
                                    "1 - Cardiologist\n"
                                    "2 - Dermatologist\n"
                                    "3 - Gynecologist\n"
                                    "4 - Neurologist\n"
                                    "5 - Pediatrician\n"
                                    "6 - Exit\n"))

            if specialist == 6:
                return  

            clear_screen()

            desired_date = input('Enter the desired date (DD/MM/YYYY): ')

            file_path = pathlib.Path("users.xlsx")

            if file_path.exists():
                workbook = openpyxl.load_workbook("users.xlsx")
            else:
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet['D1'] = "Appointments"

            sheet = workbook.active

            user_email = input("Enter the user's email to schedule the appointment: ")
            user_row = None

            for row in sheet.iter_rows(min_row=2, max_col=2, values_only=True):
                if row[1] == user_email: 
                    user_row = row
                    break

            if user_row:
                # Save the appointment information in the format: "Specialist date, Specialist date, ..."
                appointment_info = f"{specialist} {desired_date}"

                # Check if there's already a value in the "Appointments" column for the user
                appointments_column = sheet['D']
                user_appointments_cell = appointments_column[row[0].row - 1]  

                if user_appointments_cell.value:
                    user_appointments_cell.value += f", {appointment_info}"
                else:
                    user_appointments_cell.value = appointment_info

                workbook.save("users.xlsx")
                clear_screen()
                print('Appointment scheduled successfully')
                time.sleep(2)

            else:
                print("User not found.")

        except FileNotFoundError:
            print('File not found. Make sure the users.xlsx file exists.')
        except Exception as e:
            print(f'Error scheduling appointment: {e}')
