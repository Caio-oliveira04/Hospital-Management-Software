import openpyxl
import pathlib
import os
import time

def clear_screen():
    # Limpa a tela do terminal com base no sistema operacional
    os.system('cls' if os.name == 'nt' else 'clear')

class Usuario:
    def __init__(self):
        self.nome = ""
        self.idade = None
        self.email = ""
        self.senha = ""
        self.celular = ""

    def set_nome(self, nome):
        self.nome = nome

    def get_nome(self):
        return self.nome

    def set_idade(self, idade):
        self.idade = idade

    def get_idade(self):
        return self.idade

    def set_email(self, email):
        self.email = email

    def get_email(self):
        return self.email

    def set_senha(self, senha):
        self.senha = senha

    def get_senha(self):
        return self.senha

    def set_celular(self, celular):
        self.celular = celular

    def get_celular(self):
        return self.celular

    def cadastro_user(self, nome, email, senha):
        try: 
            arquivo_path = pathlib.Path("clientes.xlsx")

            if arquivo_path.exists():
                arquivo = openpyxl.load_workbook("clientes.xlsx")
            else:
                arquivo = openpyxl.Workbook()
                folha = arquivo.active
                folha['A1'] = "Nome"
                folha['B1'] = "Email"
                folha['C1'] = "Senha"

            folha = arquivo.active
            folha.append([nome, email, senha])

            arquivo.save("clientes.xlsx")
            clear_screen()  # Limpa a tela após o usuário fazer uma escolha
            print('Cadastro feito com sucesso')
            time.sleep(2)  # Aguarda 2 segundos

        except FileNotFoundError:
            print('Arquivo não encontrado. Certifique-se de que o arquivo clientes.xlsx existe.')

        except Exception as e:
            print(f'Erro ao salvar os dados: {e}')

    
    def login_user(self, email, senha):
        try:
            # Caminho do arquivo
            arquivo_path = pathlib.Path("clientes.xlsx")

            if arquivo_path.exists():
                # Carrega o arquivo Excel
                arquivo = openpyxl.load_workbook(arquivo_path)
                folha = arquivo.active

                # Itera sobre as linhas buscando pelo email
                for row in folha.iter_rows(min_row=2, max_col=2, values_only=True):
            
                    if row[0] == email:
                        # Se o email é encontrado, verifica a senha
                        if row[1] == senha:
                            print("Login feito com sucesso!")
                            return
                        else:
                            print("Senha inválida.")
                            return

                # Se o loop terminar sem encontrar o email
                print("Email não encontrado.")
            else:
                # Se o arquivo não existir
                print('Arquivo não encontrado. Certifique-se de que o arquivo clientes.xlsx existe.')

        except FileNotFoundError:
            print('Arquivo não encontrado. Certifique-se de que o arquivo clientes.xlsx existe.')
        except openpyxl.workbook.exceptions.PackageNotFoundError:
            print('Erro ao abrir o arquivo Excel. Verifique se o arquivo clientes.xlsx está formatado corretamente.')
        except Exception as e:
            print(f'Erro ao realizar login: {e}')
